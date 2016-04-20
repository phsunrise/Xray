from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import pickle

def get2theta(r, D):
    return np.arctan(r*1./D)/np.pi*180

wb = load_workbook("ShotSheet.xlsx")
data = pickle.load(open("data_2.pickle", 'rb'))

p1 = 43

ws = wb["LA61 shot sheet"]

# figure and axes
fig = plt.figure(figsize=(7,21))
ax1 = fig.add_subplot(3,1,1)
ax2 = fig.add_subplot(3,1,2, sharex=ax1)
ax3 = fig.add_subplot(3,1,3, sharex=ax1)
markers = ['^', 'o', 's', '*', 'h']
delaytime = {} # used to store number of instances 
               # for each delay time
shottype = 0 # AB+EF or AB or EF

for i_row in range(2, 100):
    run = ws['D'+str(i_row)].value
    try: 
        Npeaks = data[run]['Npeaks']
    except KeyError: # skip if no data exists
        continue
    
    t = ws['M'+str(i_row)].value # delay time

    if ws['F%d' % i_row].value == "Laser (AB+EF) and X-rays":
        ax = ax1
        shottype = 0
        try:
            delaytime[t] += 1
        except KeyError:
            delaytime[t] = 0
        marker = markers[delaytime[t]]
        if ws['W'+str(i_row)].value != 'Y':
            continue
    elif ws['F%d' % i_row].value == "Laser (AB) and X-rays":
        ax = ax2
        shottype = 1
        marker = 'o' 
    elif ws['F%d' % i_row].value == "Laser (EF) and X-rays":
        ax = ax3
        shottype = 2
        marker = 'o' 

    ws['X%d' % i_row] = Npeaks
    A = []
    m = []
    s = []
    if Npeaks >= 1:
        A.append(data[run]['params'][3])
        m.append(data[run]['params'][4])
        s.append(data[run]['params'][5])
        ws['Y%d' % i_row] = A[0]
        ws['Z%d' % i_row] = m[0]
        ws['AA%d' % i_row] = s[0] 
    if Npeaks >= 2:
        A.append(data[run]['params'][6])
        m.append(data[run]['params'][7])
        s.append(data[run]['params'][8])
        ws['AB%d' % i_row] = A[1]
        ws['AC%d' % i_row] = m[1]
        ws['AD%d' % i_row] = s[1] 
    if Npeaks >= 3:
        A.append(data[run]['params'][9])
        m.append(data[run]['params'][10])
        s.append(data[run]['params'][11])
        ws['AE%d' % i_row] = A[2]
        ws['AF%d' % i_row] = m[2]
        ws['AG%d' % i_row] = s[2] 

    # plotting the peak positions
    if shottype == 0: # AB+EF
        for i in xrange(Npeaks):
            if m[i] > 48.:
                c = 'b'
            elif np.abs(m[i]-p1) < 0.5:
                c = 'y'
            else:
                c = 'r'
            ax.errorbar(ws['M'+str(i_row)].value, m[i], \
                        yerr=s[i], c=c, \
                        fmt=marker)
    #END shottype 0
    else:
        for i in xrange(Npeaks):
            ax.errorbar(ws['M'+str(i_row)].value, m[i], \
                        yerr=s[i], c='r', \
                        fmt=marker)
        

    a0 = data[run]['params'][0]
    b0 = data[run]['params'][1]
    c0 = data[run]['params'][2]
    ws['AH%d' % i_row] = a0
    ws['AI%d' % i_row] = b0
    ws['AJ%d' % i_row] = c0
    #ax2.errorbar(ws['M'+str(i_row)].value, \
    #             a0*np.exp(-b0*p1)+c0, \
    #             yerr=s2, c='g', fmt='o')

ax1.axhline(y=p1, label="cold peak", \
            color='k', linestyle='--')
ax2.axhline(y=p1, label="cold peak", \
            color='k', linestyle='--')
ax3.axhline(y=p1, label="cold peak", \
            color='k', linestyle='--')
ax1.axhline(y=37.2, label="cold peak", \
            color='k', linestyle='--')
ax2.axhline(y=37.2, label="cold peak", \
            color='k', linestyle='--')
ax3.axhline(y=37.2, label="cold peak", \
            color='k', linestyle='--')
wb.save("ShotSheet.xlsx")

ax3.set_xlabel("delay (ns)")
ax1.set_ylabel(r"$2\theta$ (deg)")
ax2.set_ylabel(r"$2\theta$ (deg)")
ax3.set_ylabel(r"$2\theta$ (deg)")
ax1.set_title("AB+EF")
ax2.set_title("AB")
ax3.set_title("EF")

xmin, xmax = ax1.get_xlim()
ax1.set_xlim(xmin-0.1, xmax+0.1)
plt.show()

fig.savefig("peaks.png")
fig.savefig("peaks.pdf")
